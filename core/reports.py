# core/reports.py

from datetime import date, timedelta, datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import Participant


def generate_weekly_excel(participant_id=None):
    """
    Generate Excel file with research data - one row per week per participant.
    
    Args:
        participant_id: If provided, export only this participant. Otherwise export all.
    
    Returns:
        HttpResponse with Excel file
    """
    
    # Get participants (exclude staff and superuser)
    if participant_id:
        participants = Participant.objects.select_related('user').filter(
            id=participant_id,
            user__is_staff=False,
            user__is_superuser=False
        )
    else:
        participants = Participant.objects.select_related('user').filter(
            user__is_staff=False,
            user__is_superuser=False
        ).order_by('id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Data"
    
    # Define headers
    headers = [
        'subject_ID',
        'tx_arm',
        'start_date',
        'week_number',
        'week_start_date',
        'week_average',
        'days_with_data',
        'total_steps',
        'previous_week_target',
        'reached_goal',
        'increment',
        'new_target',
    ]
    
    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    current_row = 2
    
    # Process each participant
    for participant in participants:
        daily_steps = participant.daily_steps or []
        targets = participant.targets or {}
        
        if not daily_steps:
            continue  # Skip participants with no data
        
        # Sort daily steps by date
        sorted_steps = sorted(daily_steps, key=lambda x: x.get('date', ''))
        
        # Group steps by week
        weeks_data = {}
        
        for step_entry in sorted_steps:
            step_date = step_entry.get('date')
            step_value = step_entry.get('value', 0)
            
            if not step_date:
                continue
            
            step_date_obj = date.fromisoformat(step_date)
            days_since_start = (step_date_obj - participant.start_date).days
            
            # Determine which week this day belongs to
            week_number = (days_since_start // 7) + 1
            
            if week_number not in weeks_data:
                weeks_data[week_number] = {
                    'steps': [],
                    'week_start': participant.start_date + timedelta(days=(week_number - 1) * 7),
                    'week_end': participant.start_date + timedelta(days=(week_number - 1) * 7 + 6)
                }
            
            weeks_data[week_number]['steps'].append(step_value)
        
        # Write one row per week
        for week_num in sorted(weeks_data.keys()):
            week_info = weeks_data[week_num]
            week_start = week_info['week_start']
            week_end = week_info['week_end']
            week_steps = week_info['steps']
            
            # Calculate week statistics
            days_with_data = len(week_steps)
            total_steps = sum(week_steps)
            avg_steps = int(total_steps / days_with_data) if days_with_data > 0 else 0
            
            # Get target data - targets are keyed by the NEXT week's start date
            # because they contain THIS week's average and NEXT week's target
            next_week_start = week_start + timedelta(days=7)
            next_week_key = next_week_start.strftime("%Y-%m-%d")
            target_data = targets.get(next_week_key, {})
            
            # Get previous week's target (which is stored in THIS week's entry)
            week_key = week_start.strftime("%Y-%m-%d")
            this_week_data = targets.get(week_key, {})
            prev_target = this_week_data.get('new_target', None) if week_num > 1 else None
            
            # Determine if goal was reached
            if week_num == 1:
                reached_goal = 'NA'
            elif prev_target and target_data:
                average_from_target = target_data.get('average_steps', avg_steps)
                # Ensure both values are integers for comparison
                try:
                    avg_int = int(average_from_target)
                    target_int = int(prev_target)
                    reached_goal = 'Yes' if avg_int >= target_int else 'No'
                except (ValueError, TypeError):
                    reached_goal = 'NA'
            elif not target_data:
                reached_goal = '4'  # Not enough valid dates
            else:
                reached_goal = 'NA'
            
            # Get increment
            if target_data:
                increment = target_data.get('increase', '')
            else:
                increment = ''
            
            # Get new target
            if target_data and target_data.get('new_target'):
                new_target = target_data['new_target']
            elif week_num > 1 and prev_target:
                # Carry forward previous target if no new one calculated
                new_target = prev_target
            else:
                new_target = ''
            
            # Write row
            ws.cell(row=current_row, column=1, value=participant.user.email)
            ws.cell(row=current_row, column=2, value=participant.treatment_arm)
            
            # start_date only on first week for this participant
            if week_num == 1:
                ws.cell(row=current_row, column=3, value=participant.start_date)
            
            ws.cell(row=current_row, column=4, value=week_num)
            ws.cell(row=current_row, column=5, value=week_start)  # Show every week
            ws.cell(row=current_row, column=6, value=avg_steps)
            ws.cell(row=current_row, column=7, value=days_with_data)
            ws.cell(row=current_row, column=8, value=total_steps)
            ws.cell(row=current_row, column=9, value=prev_target if prev_target else 'NA')
            ws.cell(row=current_row, column=10, value=reached_goal)
            ws.cell(row=current_row, column=11, value=increment)
            ws.cell(row=current_row, column=12, value=new_target)
            
            current_row += 1
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create Dictionary sheet
    ws_dict = wb.create_sheet("Dictionary")
    dict_headers = ['variable', 'Plain English', 'Choices', 'Note']
    
    for col_num, header in enumerate(dict_headers, 1):
        cell = ws_dict.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Add dictionary entries
    dictionary_data = [
        ('email', 'Participant email', 'unique identifier', None),
        ('tx_arm', 'treatment arm', '0=Control | 1=Intervention', 'Does not change week to week'),
        ('start_date', 'First day the Fitbit was used', 'YYYY-MM-DD format', 'Shown only on first row per participant'),
        ('week_number', 'Study week number', '1, 2, 3, ...', 'Week 1 is baseline'),
        ('week_start_date', 'First day of this week', 'YYYY-MM-DD format', 'Shown for every week'),
        ('week_average', 'Average daily steps for this week', 'whole number', 'Rounded down'),
        ('days_with_data', 'Number of days with step data', '0-7', 'Days with synced Fitbit data'),
        ('total_steps', 'Total steps for the week', 'whole number', 'Sum of all daily steps'),
        ('previous_week_target', 'This week\'s target that was set last week', 'steps/day', 'Set at end of previous week'),
        ('reached_goal', 'Did participant reach the target?', 'Yes | No | NA | 4', 'NA for week 1; 4 = insufficient data'),
        ('increment', 'How target was adjusted', '+250, +500, +1000, maintain, etc.', 'Based on algorithm'),
        ('new_target', 'Target set for next week', 'steps/day', 'Calculated at end of this week'),
    ]
    
    for row_num, (var, plain, choices, note) in enumerate(dictionary_data, 2):
        ws_dict.cell(row=row_num, column=1, value=var)
        ws_dict.cell(row=row_num, column=2, value=plain)
        ws_dict.cell(row=row_num, column=3, value=choices)
        ws_dict.cell(row=row_num, column=4, value=note)
    
    # Auto-size dictionary columns
    for col in range(1, len(dict_headers) + 1):
        ws_dict.column_dimensions[get_column_letter(col)].width = 30
    
    # Prepare response
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if participant_id:
        filename = f'partnersteps_weekly_participant_{participant_id}_{timestamp}.xlsx'
    else:
        filename = f'partnersteps_weekly_summary_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def generate_daily_excel(participant_id=None):
    """
    Generate Excel file with daily step data - one row per day.
    
    Args:
        participant_id: If provided, export only this participant. Otherwise export all.
    
    Returns:
        HttpResponse with Excel file
    """
    
    # Get participants (exclude staff and superuser)
    if participant_id:
        participants = Participant.objects.select_related('user').filter(
            id=participant_id,
            user__is_staff=False,
            user__is_superuser=False
        )
    else:
        participants = Participant.objects.select_related('user').filter(
            user__is_staff=False,
            user__is_superuser=False
        ).order_by('id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Data"
    
    # Define headers
    headers = [
        'email',
        'tx_arm',
        'start_date',
        'date',
        'day_number',
        'week_number',
        'daily_steps',
        'week_total',
        'week_average',
        'previous_week_target',
        'reached_goal',
        'increment',
        'new_target',
    ]
    
    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    current_row = 2
    
    # Process each participant
    for participant in participants:
        daily_steps = participant.daily_steps or []
        targets = participant.targets or {}
        
        if not daily_steps:
            continue
        
        # Sort daily steps by date
        sorted_steps = sorted(daily_steps, key=lambda x: x.get('date', ''))
        
        # Track for week calculations
        week_steps = []
        current_week = None
        first_row_for_participant = True
        
        for step_entry in sorted_steps:
            step_date = step_entry.get('date')
            step_value = step_entry.get('value', 0)
            
            if not step_date:
                continue
            
            step_date_obj = date.fromisoformat(step_date)
            days_since_start = (step_date_obj - participant.start_date).days
            day_number = days_since_start + 1
            week_number = (days_since_start // 7) + 1
            
            # Check if we're starting a new week
            if current_week != week_number:
                week_steps = []
                current_week = week_number
            
            week_steps.append(step_value)
            
            # Write basic daily data
            ws.cell(row=current_row, column=1, value=participant.user.email)
            ws.cell(row=current_row, column=2, value=participant.treatment_arm)
            
            # start_date only on first row
            if first_row_for_participant:
                ws.cell(row=current_row, column=3, value=participant.start_date)
                first_row_for_participant = False
            
            ws.cell(row=current_row, column=4, value=step_date_obj)
            ws.cell(row=current_row, column=5, value=day_number)
            ws.cell(row=current_row, column=6, value=week_number)
            ws.cell(row=current_row, column=7, value=step_value)
            
            # Check if this is last day of week (day 7, 14, 21, etc.)
            if day_number % 7 == 0 and day_number >= 7:
                # Calculate week totals
                week_total = sum(week_steps)
                week_avg = int(week_total / len(week_steps)) if week_steps else 0
                
                ws.cell(row=current_row, column=8, value=week_total)
                ws.cell(row=current_row, column=9, value=week_avg)
                
                # Get target data - targets are keyed by the NEXT week's start date
                week_start_date = participant.start_date + timedelta(days=(week_number - 1) * 7)
                next_week_start = week_start_date + timedelta(days=7)
                next_week_key = next_week_start.strftime("%Y-%m-%d")
                target_data = targets.get(next_week_key, {})
                
                # Get previous week's target (stored in THIS week's entry)
                week_key = week_start_date.strftime("%Y-%m-%d")
                this_week_data = targets.get(week_key, {})
                prev_target = this_week_data.get('new_target', None) if week_number > 1 else None
                ws.cell(row=current_row, column=10, value=prev_target if prev_target else 'NA')
                
                # Reached goal
                if week_number == 1:
                    reached_goal = 'NA'
                elif prev_target and target_data:
                    average_from_target = target_data.get('average_steps', week_avg)
                    try:
                        avg_int = int(average_from_target)
                        target_int = int(prev_target)
                        reached_goal = 'Yes' if avg_int >= target_int else 'No'
                    except (ValueError, TypeError):
                        reached_goal = 'NA'
                elif not target_data:
                    reached_goal = '4'
                else:
                    reached_goal = 'NA'
                
                ws.cell(row=current_row, column=11, value=reached_goal)
                
                # Increment and new target
                if target_data:
                    ws.cell(row=current_row, column=12, value=target_data.get('increase', ''))
                    ws.cell(row=current_row, column=13, value=target_data.get('new_target', ''))
                elif week_number > 1 and prev_target:
                    ws.cell(row=current_row, column=12, value='')
                    ws.cell(row=current_row, column=13, value=prev_target)
            
            current_row += 1
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Create Dictionary sheet
    ws_dict = wb.create_sheet("Dictionary")
    dict_headers = ['variable', 'Plain English', 'Choices', 'Note']
    
    for col_num, header in enumerate(dict_headers, 1):
        cell = ws_dict.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    dictionary_data = [
        ('email', 'Participant email', 'unique identifier', None),
        ('tx_arm', 'treatment arm', '0=Control | 1=Intervention', None),
        ('start_date', 'First day Fitbit was used', 'YYYY-MM-DD', 'Shown on first row only'),
        ('date', 'Date for this row', 'YYYY-MM-DD', 'Daily date'),
        ('day_number', 'Day number in study', '1, 2, 3, ...', 'Days since start'),
        ('week_number', 'Week number in study', '1, 2, 3, ...', 'Week 1 is baseline'),
        ('daily_steps', 'Steps for this day', 'whole number', None),
        ('week_total', 'Total steps for week', 'whole number', 'Shown on last day of week'),
        ('week_average', 'Average steps for week', 'whole number', 'Shown on last day of week'),
        ('previous_week_target', 'This week\'s target that was set last week', 'steps/day', 'Set at end of previous week'),
        ('reached_goal', 'Met weekly target?', 'Yes | No | NA | 4', 'Shown on last day of week'),
        ('increment', 'Target adjustment', '+250, +500, +1000, maintain', 'Shown on last day of week'),
        ('new_target', 'Target for next week', 'steps/day', 'Shown on last day of week'),
    ]
    
    for row_num, (var, plain, choices, note) in enumerate(dictionary_data, 2):
        ws_dict.cell(row=row_num, column=1, value=var)
        ws_dict.cell(row=row_num, column=2, value=plain)
        ws_dict.cell(row=row_num, column=3, value=choices)
        ws_dict.cell(row=row_num, column=4, value=note)
    
    for col in range(1, len(dict_headers) + 1):
        ws_dict.column_dimensions[get_column_letter(col)].width = 30
    
    # Prepare response
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if participant_id:
        filename = f'partnersteps_daily_participant_{participant_id}_{timestamp}.xlsx'
    else:
        filename = f'partnersteps_daily_data_{timestamp}.xlsx'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response